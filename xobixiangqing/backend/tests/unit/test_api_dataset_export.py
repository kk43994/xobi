"""
Dataset 导出（ExportProfile v1）单测：按 taiyang.xlsx 覆盖导出并追加 image1..imageN 列
"""

import io

import openpyxl

from conftest import assert_success_response


def _make_taiyang_xlsx() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "SKUID",
        "产品名称",
        "产品分类",
        "原价格",
        "折扣价",
        "产品图片",
        "包裹重量",
        "包裹尺寸",
        "SKU名称",
        "平台SKU",
        "SKU图片",
        "产品ID",
    ]
    ws.append(headers)
    ws.append(
        [
            "\t430087286034",
            "旧标题",
            "Home & Living>Dinnerware>Cutleries",
            40,
            20,
            "http://old1.png,http://old2.png",
            0.05,
            "长：10,宽：10, 高：5",
            "SKU-橙色",
            "N002",
            "http://old-sku.png",
            "\t44800974685",
        ]
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_export_dataset_excel_overwrite_with_image_columns(client):
    create = client.post(
        "/api/datasets/create-from-excel",
        data={
            "template_key": "taiyang",
            "name": "test-dataset",
            "file": (_make_taiyang_xlsx(), "taiyang.xlsx"),
        },
        content_type="multipart/form-data",
    )
    data = assert_success_response(create)
    dataset_id = data["data"]["dataset"]["id"]
    item_id = data["data"]["preview_items"][0]["id"]

    patch = client.patch(
        f"/api/datasets/{dataset_id}/items/{item_id}",
        json={
            "new_title": "新标题",
            "new_images": ["http://new1.png", "http://new2.png"],
            "status": "done",
        },
    )
    assert_success_response(patch)

    export = client.post(
        f"/api/datasets/{dataset_id}/export-excel",
        json={"mode": "overwrite", "image_columns": True, "max_images": 3},
    )
    data = assert_success_response(export)
    download_url = data["data"]["download_url"]
    assert isinstance(download_url, str) and download_url.startswith("/api/assets/")

    dl = client.get(download_url)
    assert dl.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(dl.data))
    ws = wb.active

    header = [c.value for c in ws[1]]
    assert "产品名称" in header
    assert "产品图片" in header
    assert "SKU图片" in header
    assert "image1" in header
    assert "image3" in header

    row2 = [c.value for c in ws[2]]
    m = dict(zip(header, row2))

    assert m["产品名称"] == "新标题"
    assert m["产品图片"] == "http://new1.png,http://new2.png"
    assert m["SKU图片"] == "http://new1.png"
    assert m["image1"] == "http://new1.png"
    assert m["image2"] == "http://new2.png"
    assert m["image3"] in (None, "")


def test_create_projects_from_dataset_sets_project_id(client):
    create = client.post(
        "/api/datasets/create-from-excel",
        data={
            "template_key": "taiyang",
            "name": "test-dataset",
            "file": (_make_taiyang_xlsx(), "taiyang.xlsx"),
        },
        content_type="multipart/form-data",
    )
    data = assert_success_response(create)
    dataset_id = data["data"]["dataset"]["id"]
    item_id = data["data"]["preview_items"][0]["id"]

    create_proj = client.post(
        f"/api/datasets/{dataset_id}/projects/create",
        json={"item_ids": [item_id], "download_material": False, "platform_hint": "shopee"},
    )
    data = assert_success_response(create_proj)
    assert data["data"]["created"] == 1
    project_id = data["data"]["results"][0]["project_id"]
    assert isinstance(project_id, str) and project_id

    items = client.get(f"/api/datasets/{dataset_id}/items", query_string={"limit": 50, "offset": 0})
    data = assert_success_response(items)
    row = next((x for x in data["data"]["items"] if x["id"] == item_id), None)
    assert row is not None
    assert row["project_id"] == project_id

    proj = client.get(f"/api/projects/{project_id}")
    data = assert_success_response(proj)
    assert data["data"]["project_id"] == project_id


def test_download_dataset_template_taiyang(client):
    res = client.get("/api/datasets/templates/taiyang")
    assert res.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(res.data))
    ws = wb.active
    header = [c.value for c in ws[1]]

    assert header[:3] == ["SKUID", "产品名称", "产品分类"]
    assert "产品图片" in header
    assert "SKU图片" in header
    assert "产品ID" in header
