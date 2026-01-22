"""
Dataset Controller (Phase 2)

Dataset = one imported Excel/CSV file
DatasetItem = one row/SKU

Minimal API:
- POST /api/datasets/create-from-excel
- GET  /api/datasets
- GET  /api/datasets/<dataset_id>
- GET  /api/datasets/<dataset_id>/items
- PATCH /api/datasets/<dataset_id>/items/<item_id>

Phase 2 bridge (MVP): create B style batch job from dataset items
- POST /api/datasets/<dataset_id>/jobs/style-batch
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from flask import Blueprint, current_app, request, send_file
from werkzeug.utils import secure_filename

from models import Asset, Dataset, DatasetItem, Job, Material, Project, db
from services import get_file_service
from services.dataset_jobs import start_title_rewrite_job
from services.job_sync import start_auto_sync_b_style_batch_job
from services.legacy_b_client import create_style_batch_from_items
from utils import error_response, success_response

logger = logging.getLogger(__name__)

dataset_bp = Blueprint("datasets", __name__, url_prefix="/api/datasets")


def _safe_int(value: Optional[str], default: int, *, min_value: int, max_value: int) -> int:
    try:
        n = int(str(value).strip())
    except Exception:
        return default
    return max(min_value, min(max_value, n))


def _detect_kind_by_filename(filename: str) -> str:
    ext = Path(filename or "").suffix.lower()
    if ext in (".xlsx", ".xls", ".csv"):
        return "excel"
    return "file"


def _split_images(value: Any) -> List[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    s = s.replace("，", ",")
    parts = re.split(r"[,\n]+", s)
    out: List[str] = []
    for p in parts:
        v = str(p).strip()
        if not v:
            continue
        out.append(v)
    # de-dup while preserving order
    seen = set()
    uniq: List[str] = []
    for u in out:
        if u in seen:
            continue
        seen.add(u)
        uniq.append(u)
    return uniq


def _strip_excel_id(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    # Some sheets prefix a tab to prevent scientific notation in Excel.
    return s.lstrip("\t").strip()


def _save_uploaded_asset(file_storage, *, system: str = "A") -> Tuple[Asset, Path]:
    original_name = str(file_storage.filename or "")
    safe_name = secure_filename(original_name) or "file"
    kind = _detect_kind_by_filename(safe_name)

    asset = Asset(system=system, kind=kind, name=original_name, storage="local")
    db.session.add(asset)
    db.session.flush()

    upload_root = Path(current_app.config["UPLOAD_FOLDER"]).resolve()
    asset_dir = (upload_root / "assets" / asset.id).resolve()
    asset_dir.mkdir(parents=True, exist_ok=True)

    file_path = (asset_dir / safe_name).resolve()
    file_storage.save(str(file_path))

    asset.file_path = file_path.relative_to(upload_root).as_posix()
    asset.content_type = getattr(file_storage, "mimetype", None)
    try:
        asset.size_bytes = int(file_path.stat().st_size)
    except Exception:
        asset.size_bytes = None
    asset.set_meta({"original_filename": original_name})
    return asset, file_path


def _read_table_from_csv(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    # Best-effort encoding handling for Excel-exported CSV.
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                headers = [h.strip() for h in (reader.fieldnames or []) if str(h or "").strip()]
                rows: List[Dict[str, Any]] = []
                for row in reader:
                    if not isinstance(row, dict):
                        continue
                    rows.append({k: row.get(k) for k in headers})
                return headers, rows
        except Exception:
            continue
    raise RuntimeError("Failed to read CSV (unsupported encoding)")


def _read_table_from_excel(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    ext = path.suffix.lower()

    if ext == ".csv":
        return _read_table_from_csv(path)

    if ext == ".xls":
        import xlrd  # type: ignore

        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        headers = [h for h in headers if h]
        rows: List[Dict[str, Any]] = []
        for r in range(1, sheet.nrows):
            data: Dict[str, Any] = {}
            for c, h in enumerate(headers):
                if c >= sheet.ncols:
                    continue
                data[h] = sheet.cell_value(r, c)
            rows.append(data)
        return headers, rows

    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    first = next(it, None)
    if not first:
        return [], []
    headers = [str(x).strip() if x is not None else "" for x in first]
    headers = [h for h in headers if h]

    rows: List[Dict[str, Any]] = []
    for row in it:
        data: Dict[str, Any] = {}
        for idx, h in enumerate(headers):
            data[h] = row[idx] if idx < len(row) else None
        # skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in data.values()):
            continue
        rows.append(data)
    return headers, rows


def _build_dataset_items_from_rows(rows: List[Dict[str, Any]]) -> List[DatasetItem]:
    items: List[DatasetItem] = []
    # DatasetItem.row_index: use Excel-like row index (header row = 1, first data row = 2)
    excel_row_index = 2
    for r in rows:
        skuid = _strip_excel_id(r.get("SKUID"))
        product_id = _strip_excel_id(r.get("产品ID"))
        platform_sku = _strip_excel_id(r.get("平台SKU"))

        title = str(r.get("产品名称") or "").strip()
        category_path = str(r.get("产品分类") or "").strip()

        raw_images = r.get("产品图片")
        images = _split_images(raw_images)

        variant_name = str(r.get("SKU名称") or "").strip()
        variant_image = str(r.get("SKU图片") or "").strip()

        original_price = r.get("原价格")
        discount_price = r.get("折扣价")

        weight = r.get("包裹重量")
        size_raw = str(r.get("包裹尺寸") or "").strip()

        item = DatasetItem(dataset_id="__tmp__", row_index=excel_row_index)
        item.set_external_ids(
            {
                "skuid": skuid,
                "product_id": product_id,
                "platform_sku": platform_sku,
            }
        )
        item.title = title
        item.category_path = category_path
        item.set_images(images)
        item.variant_name = variant_name
        item.variant_image = variant_image
        item.set_price({"original": original_price, "discount": discount_price})
        item.set_package({"weight": weight, "size_raw": size_raw})
        item.set_attributes({"raw": r})
        item.status = "pending"
        item.set_errors([])
        item.set_asset_ids([])

        items.append(item)
        excel_row_index += 1

    return items


def _taiyang_default_columns() -> List[str]:
    return [
        "SKUID",
        "产品名称",
        "产品分类",
        "原价格",
        "折扣价",
        "产品图片",
        "包裹重量",
        "包裹尺寸",
        "SKU名称",
        "平台SKU",
        "SKU图片",
        "产品ID",
    ]


@dataset_bp.route("/templates/<template_key>", methods=["GET"])
def download_dataset_template(template_key: str):
    """
    GET /api/datasets/templates/<template_key>

    v1: 只提供 taiyang 模板（与 `taiyang.xlsx` 列结构一致），方便用户下载后批量导入。
    """
    key = (template_key or "").strip().lower()
    for ext in (".xlsx", ".xls", ".csv"):
        if key.endswith(ext):
            key = key[: -len(ext)]
            break

    if key != "taiyang":
        return error_response("TEMPLATE_NOT_FOUND", "Template not found", 404)

    try:
        import openpyxl  # type: ignore

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(_taiyang_default_columns())
        ws.append(
            [
                "\t430087286034",
                "示例：产品名称（可改成你的标题）",
                "Home & Living>Dinnerware>Cutleries",
                40,
                20,
                "https://example.com/image1.png,https://example.com/image2.png",
                0.05,
                "长：10,宽：10,高：5",
                "SKU名称示例",
                "平台SKU示例",
                "https://example.com/sku_image.png",
                "\t44800974685",
            ]
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="taiyang.xlsx",
        )
    except Exception as e:
        logger.error("download_dataset_template failed: %s", e, exc_info=True)
        return error_response("SERVER_ERROR", str(e), 500)


def _build_project_idea_prompt_from_item(item: DatasetItem, *, platform_hint: Optional[str] = None) -> str:
    title = (item.new_title or "").strip() or (item.title or "").strip() or "电商产品"
    category = (item.category_path or "").strip()
    sku_name = (item.variant_name or "").strip()
    ext = item.get_external_ids() or {}
    price = item.get_price() or {}
    package = item.get_package() or {}

    lines = [
        "请为跨境电商生成一套详情页图片图集（建议 6-8 张，风格统一，可直接上架使用）。",
        f"平台偏好：{platform_hint}" if platform_hint else "平台偏好：通用电商",
        "",
        f"产品名称：{title}",
    ]
    if category:
        lines.append(f"产品分类：{category}")
    if sku_name:
        lines.append(f"SKU名称：{sku_name}")

    if ext.get("skuid"):
        lines.append(f"SKUID：{ext.get('skuid')}")
    if ext.get("platform_sku"):
        lines.append(f"平台SKU：{ext.get('platform_sku')}")
    if ext.get("product_id"):
        lines.append(f"产品ID：{ext.get('product_id')}")

    if price:
        original = price.get("original")
        discount = price.get("discount")
        if original is not None or discount is not None:
            lines.append(f"价格：原价 {original} / 折扣价 {discount}")

    if package:
        weight = package.get("weight")
        size_raw = package.get("size_raw")
        if weight is not None or size_raw:
            lines.append(f"包装：重量 {weight} / 尺寸 {size_raw}")

    lines.extend(
        [
            "",
            "要求：",
            "- 画面清晰、质感高级、构图简洁，符合电商审美",
            "- 卖点表达真实具体，避免夸张虚假描述",
            "- 适合移动端浏览，文字少而大，信息层级清晰",
        ]
    )

    return "\n".join([x for x in lines if x is not None])


@dataset_bp.route("/create-from-excel", methods=["POST"])
def create_from_excel():
    """
    POST /api/datasets/create-from-excel

    Accepts:
    - multipart/form-data with file=...
    or
    - JSON with {asset_id: "..."} pointing to a previously uploaded excel asset (local storage).
    """
    try:
        template_key = (request.form.get("template_key") or request.args.get("template_key") or "taiyang").strip() or "taiyang"
        name_override = (request.form.get("name") or request.args.get("name") or "").strip()

        asset: Optional[Asset] = None
        local_path: Optional[Path] = None

        if "file" in request.files:
            f = request.files["file"]
            if not f or not getattr(f, "filename", ""):
                return error_response("INVALID_REQUEST", "Invalid file", 400)
            asset, local_path = _save_uploaded_asset(f, system="A")
        else:
            payload = request.get_json(silent=True) or {}
            asset_id = str(payload.get("asset_id") or "").strip()
            if not asset_id:
                return error_response("INVALID_REQUEST", "Missing file or asset_id", 400)
            asset = Asset.query.get(asset_id)
            if not asset or (asset.storage or "").lower() != "local":
                return error_response("ASSET_NOT_FOUND", "Excel asset not found", 404)
            upload_root = Path(current_app.config["UPLOAD_FOLDER"]).resolve()
            rel = str(asset.file_path or "").replace("\\", "/").lstrip("/")
            local_path = (upload_root / rel).resolve()
            if not local_path.exists():
                return error_response("ASSET_NOT_FOUND", "Excel file not found", 404)

        assert asset is not None and local_path is not None

        headers, rows = _read_table_from_excel(local_path)
        if not headers:
            return error_response("INVALID_EXCEL", "Empty sheet or missing header row", 400)

        dataset_name = name_override or (asset.name or local_path.name)
        dataset = Dataset(name=dataset_name, template_key=template_key, status="active", source_asset_id=asset.id)
        dataset.set_columns(headers)
        dataset.set_mapping({"v1": "taiyang", "note": "auto-mapped by known column names"})
        db.session.add(dataset)
        db.session.flush()

        # Build items
        items = _build_dataset_items_from_rows(rows)
        for it in items:
            it.dataset_id = dataset.id
            db.session.add(it)

        # Record an import job (optional but useful for traceability)
        job = Job(system="A", job_type="IMPORT_EXCEL", status="succeeded", dataset_id=dataset.id)
        job.set_progress({"total": len(items), "completed": len(items), "failed": 0})
        job.set_meta({"source_asset_id": asset.id, "template_key": template_key})
        job.completed_at = datetime.utcnow()
        db.session.add(job)

        db.session.commit()

        preview = [it.to_dict() for it in items[:10]]
        return success_response(
            {"dataset": dataset.to_dict(include_counts=True), "preview_items": preview, "import_job_id": job.id},
            message="dataset_created",
        )
    except Exception as e:
        logger.error("create_from_excel failed: %s", e, exc_info=True)
        db.session.rollback()
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/", methods=["GET"], strict_slashes=False)
def list_datasets():
    try:
        limit = _safe_int(request.args.get("limit"), 50, min_value=1, max_value=200)
        offset = _safe_int(request.args.get("offset"), 0, min_value=0, max_value=1000000)

        q = Dataset.query.order_by(Dataset.created_at.desc())
        total = q.count()
        rows = q.offset(offset).limit(limit).all()
        return success_response({"datasets": [d.to_dict(include_counts=True) for d in rows], "total": total})
    except Exception as e:
        logger.error("list_datasets failed: %s", e, exc_info=True)
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>", methods=["GET"])
def get_dataset(dataset_id: str):
    try:
        dataset_id = (dataset_id or "").strip()
        ds = Dataset.query.get(dataset_id)
        if not ds:
            return error_response("DATASET_NOT_FOUND", "Dataset not found", 404)
        return success_response({"dataset": ds.to_dict(include_counts=True)})
    except Exception as e:
        logger.error("get_dataset failed: %s", e, exc_info=True)
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>/items", methods=["GET"])
def list_dataset_items(dataset_id: str):
    try:
        dataset_id = (dataset_id or "").strip()
        ds = Dataset.query.get(dataset_id)
        if not ds:
            return error_response("DATASET_NOT_FOUND", "Dataset not found", 404)

        limit = _safe_int(request.args.get("limit"), 50, min_value=1, max_value=200)
        offset = _safe_int(request.args.get("offset"), 0, min_value=0, max_value=1000000)
        status = (request.args.get("status") or "").strip().lower() or None
        keyword = (request.args.get("q") or "").strip()

        q = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id)
        if status:
            q = q.filter(DatasetItem.status == status)
        if keyword:
            like = f"%{keyword}%"
            q = q.filter((DatasetItem.title.ilike(like)) | (DatasetItem.external_ids.ilike(like)))

        total = q.count()
        rows = q.order_by(DatasetItem.row_index.asc()).offset(offset).limit(limit).all()
        return success_response({"items": [it.to_dict() for it in rows], "total": total})
    except Exception as e:
        logger.error("list_dataset_items failed: %s", e, exc_info=True)
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>/items/<item_id>", methods=["PATCH"])
def update_dataset_item(dataset_id: str, item_id: str):
    try:
        dataset_id = (dataset_id or "").strip()
        item_id = (item_id or "").strip()

        row = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id, DatasetItem.id == item_id).first()
        if not row:
            return error_response("DATASET_ITEM_NOT_FOUND", "Dataset item not found", 404)

        payload = request.get_json(silent=True) or {}
        if not isinstance(payload, dict):
            return error_response("INVALID_REQUEST", "Invalid JSON body", 400)

        if "new_title" in payload:
            row.new_title = str(payload.get("new_title") or "").strip()
        if "new_images" in payload:
            v = payload.get("new_images")
            if isinstance(v, list):
                row.set_new_images([str(x).strip() for x in v if str(x).strip()])
        if "status" in payload:
            s = str(payload.get("status") or "").strip().lower()
            if s in ("pending", "processing", "done", "failed"):
                row.status = s
        if "errors" in payload:
            v = payload.get("errors")
            if isinstance(v, list):
                row.set_errors(v)
        if "asset_ids" in payload:
            v = payload.get("asset_ids")
            if isinstance(v, list):
                row.set_asset_ids([str(x).strip() for x in v if str(x).strip()])

        db.session.commit()
        return success_response({"item": row.to_dict()}, message="updated")
    except Exception as e:
        logger.error("update_dataset_item failed: %s", e, exc_info=True)
        db.session.rollback()
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>/items/<item_id>", methods=["GET"])
def get_dataset_item(dataset_id: str, item_id: str):
    try:
        dataset_id = (dataset_id or "").strip()
        item_id = (item_id or "").strip()

        row = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id, DatasetItem.id == item_id).first()
        if not row:
            return error_response("DATASET_ITEM_NOT_FOUND", "Dataset item not found", 404)

        return success_response({"item": row.to_dict()})
    except Exception as e:
        logger.error("get_dataset_item failed: %s", e, exc_info=True)
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>/items/<item_id>/project/create", methods=["POST"])
def create_project_from_dataset_item(dataset_id: str, item_id: str):
    """
    POST /api/datasets/<dataset_id>/items/<item_id>/project/create

    Create (or open existing) Project for one DatasetItem.

    Body (JSON, optional):
      - platform_hint?: string
      - download_material?: true|false (default true)
      - material_url?: string  # preferred material image url; defaults to new_images[0] then variant_image then images[0]
      - force_new?: true|false (default false)  # create a new project even if item.project_id already exists
    """
    dataset_id = (dataset_id or "").strip()
    item_id = (item_id or "").strip()

    ds = Dataset.query.get(dataset_id)
    if not ds:
        return error_response("DATASET_NOT_FOUND", "Dataset not found", 404)

    item = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id, DatasetItem.id == item_id).first()
    if not item:
        return error_response("DATASET_ITEM_NOT_FOUND", "Dataset item not found", 404)

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        payload = {}

    platform_hint = (str(payload.get("platform_hint") or "").strip() or None)
    download_material = payload.get("download_material")
    if download_material is None:
        download_material = True
    download_material = bool(download_material)

    force_new = bool(payload.get("force_new") or False)

    existing_project_id = (item.project_id or "").strip() or None
    if not force_new and existing_project_id and Project.query.get(existing_project_id):
        return success_response({"project_id": existing_project_id, "created": False, "item": item.to_dict()})

    material_url = str(payload.get("material_url") or "").strip() or None
    if not material_url:
        new_images = item.get_new_images()
        if new_images:
            material_url = new_images[0]

    if not material_url:
        material_url = (item.variant_image or "").strip() or None

    if not material_url:
        images = item.get_images()
        if images:
            material_url = images[0]

    file_service = get_file_service(current_app.config["UPLOAD_FOLDER"])

    def _download_image_as_pil(url: str):
        if not url:
            return None
        try:
            import httpx
            from PIL import Image  # type: ignore
            from io import BytesIO

            with httpx.Client(timeout=20.0, follow_redirects=True) as client:
                res = client.get(url, headers={"User-Agent": "Mozilla/5.0"})
                res.raise_for_status()
                content = res.content

            if not content:
                return None
            if len(content) > 25 * 1024 * 1024:
                return None

            img = Image.open(BytesIO(content))
            img.load()
            return img
        except Exception:
            return None

    try:
        idea_prompt = _build_project_idea_prompt_from_item(item, platform_hint=platform_hint)
        project = Project(creation_type="idea", project_type="ecom", idea_prompt=idea_prompt, status="DRAFT")
        db.session.add(project)
        db.session.flush()

        item.project_id = project.id

        if download_material and material_url:
            img = _download_image_as_pil(material_url)
            if img is not None:
                rel = file_service.save_material_image(img, project.id)
                filename = rel.split("/")[-1]
                url = file_service.get_file_url(project.id, "materials", filename)
                material = Material(project_id=project.id, filename=filename, relative_path=rel, url=url)
                db.session.add(material)

        db.session.commit()
        return success_response(
            {
                "project_id": project.id,
                "created": True,
                "item": item.to_dict(),
                "material_url": material_url,
            },
            message="project_created",
        )
    except Exception as e:
        logger.error("create_project_from_dataset_item failed: %s", e, exc_info=True)
        db.session.rollback()
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>/jobs/style-batch", methods=["POST"])
def create_style_batch_job(dataset_id: str):
    """
    Create a legacy B STYLE_BATCH job from dataset items, and register it into core Job table.
    """
    try:
        dataset_id = (dataset_id or "").strip()
        ds = Dataset.query.get(dataset_id)
        if not ds:
            return error_response("DATASET_NOT_FOUND", "Dataset not found", 404)

        payload = request.get_json(silent=True) or {}
        if not isinstance(payload, dict):
            payload = {}

        item_ids = payload.get("item_ids")
        if isinstance(item_ids, list) and item_ids:
            selected_ids = [str(x).strip() for x in item_ids if str(x).strip()]
            q = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id, DatasetItem.id.in_(selected_ids))
        else:
            q = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id)

        rows = q.order_by(DatasetItem.row_index.asc()).all()
        if not rows:
            return error_response("DATASET_ITEMS_NOT_FOUND", "No dataset items", 404)

        style_preset = str(payload.get("style_preset") or "shein").strip() or "shein"
        options = payload.get("options") if isinstance(payload.get("options"), dict) else {}
        requirements = str(payload.get("requirements") or "").strip()
        target_language = str(payload.get("target_language") or "same").strip() or "same"
        aspect_ratio = str(payload.get("aspect_ratio") or "1:1").strip() or "1:1"

        items: List[dict] = []
        valid_item_ids: List[str] = []

        for r in rows:
            images = r.get_images()
            image_url = (r.variant_image or "").strip() or (images[0] if images else "")
            if not image_url:
                r.status = "failed"
                errors = r.get_errors()
                msg = "缺少可用图片链接（variant_image / 产品图片）"
                if msg not in errors:
                    errors.append(msg)
                    r.set_errors(errors)
                continue

            items.append({"id": r.id, "image_url": image_url, "title": (r.title or "").strip(), "subtitle": ""})
            valid_item_ids.append(r.id)
            r.status = "processing"

        if not items:
            db.session.commit()
            return error_response("INVALID_DATASET_ITEMS", "No usable image_url for style batch", 400)

        result = create_style_batch_from_items(
            items=items,
            style_preset=style_preset,
            options=options,
            requirements=requirements,
            target_language=target_language,
            aspect_ratio=aspect_ratio,
            auto_start=True,
        )
        b_job_id = str(result.get("job_id") or "").strip()
        if not b_job_id:
            db.session.rollback()
            return error_response("LEGACY_B_ERROR", f"Failed to create B job: {result}", 502)

        job = Job(system="B", job_type="STYLE_BATCH", status="running", dataset_id=dataset_id, external_id=b_job_id)
        job.set_progress({"total": len(items), "completed": 0, "failed": 0})
        job.set_meta(
            {
                "dataset_id": dataset_id,
                "item_ids": valid_item_ids,
                "b_request": {
                    "style_preset": style_preset,
                    "options": options,
                    "requirements": requirements,
                    "target_language": target_language,
                    "aspect_ratio": aspect_ratio,
                },
            }
        )
        db.session.add(job)

        db.session.commit()

        # Auto-sync B outputs back to core Asset/DatasetItem (best-effort)
        app = current_app._get_current_object()
        start_auto_sync_b_style_batch_job(job_id=job.id, app=app)

        return success_response(
            {
                "job": job.to_dict(),
                "external_id": b_job_id,
                "message": result.get("message"),
                "preview": result.get("preview"),
            },
            message="job_created",
        )
    except Exception as e:
        logger.error("create_style_batch_job failed: %s", e, exc_info=True)
        db.session.rollback()
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>/jobs/title-rewrite", methods=["POST"])
def create_title_rewrite_job(dataset_id: str):
    """
    Create a TITLE_REWRITE_BATCH job (runs in core A, calls legacy B /api/title/rewrite per row).
    """
    try:
        dataset_id = (dataset_id or "").strip()
        ds = Dataset.query.get(dataset_id)
        if not ds:
            return error_response("DATASET_NOT_FOUND", "Dataset not found", 404)

        payload = request.get_json(silent=True) or {}
        if not isinstance(payload, dict):
            payload = {}

        item_ids = payload.get("item_ids")
        if isinstance(item_ids, list) and item_ids:
            selected_ids = [str(x).strip() for x in item_ids if str(x).strip()]
            q = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id, DatasetItem.id.in_(selected_ids))
        else:
            q = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id)

        rows = q.order_by(DatasetItem.row_index.asc()).all()
        if not rows:
            return error_response("DATASET_ITEMS_NOT_FOUND", "No dataset items", 404)

        language = str(payload.get("language") or "auto").strip() or "auto"  # auto|zh|th|en
        style = str(payload.get("style") or "simple").strip() or "simple"
        requirements = str(payload.get("requirements") or "").strip()
        max_length = int(payload.get("max_length") or 100)
        if max_length <= 0:
            max_length = 100

        selected_item_ids = [r.id for r in rows]

        job = Job(system="A", job_type="TITLE_REWRITE_BATCH", status="pending", dataset_id=dataset_id)
        job.set_progress({"total": len(selected_item_ids), "completed": 0, "failed": 0})
        job.set_meta(
            {
                "dataset_id": dataset_id,
                "item_ids": selected_item_ids,
                "params": {
                    "language": language,
                    "style": style,
                    "requirements": requirements,
                    "max_length": max_length,
                },
            }
        )
        db.session.add(job)
        db.session.commit()

        app = current_app._get_current_object()
        start_title_rewrite_job(
            job_id=job.id,
            dataset_id=dataset_id,
            item_ids=selected_item_ids,
            language=language,
            style=style,
            requirements=requirements,
            max_length=max_length,
            app=app,
        )

        return success_response({"job": job.to_dict()}, message="job_created")
    except Exception as e:
        logger.error("create_title_rewrite_job failed: %s", e, exc_info=True)
        db.session.rollback()
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>/export-excel", methods=["POST"])
def export_dataset_excel(dataset_id: str):
    """
    POST /api/datasets/<dataset_id>/export-excel

    v1: 先按 `taiyang.xlsx` 的列结构导出（同名列 overwrite），并把导出文件注册为 Asset（kind=excel）。

    Body (JSON, optional):
      - mode: overwrite|append (default overwrite)
      - image_columns: true|false (default false)  # 额外追加 image1..imageN 列
      - max_images: 1..20 (default 9)
    """
    dataset_id = (dataset_id or "").strip()
    ds = Dataset.query.get(dataset_id)
    if not ds:
        return error_response("DATASET_NOT_FOUND", "Dataset not found", 404)

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        payload = {}

    mode = str(payload.get("mode") or "overwrite").strip().lower() or "overwrite"
    if mode not in ("overwrite", "append"):
        mode = "overwrite"

    image_columns = bool(payload.get("image_columns") or False)
    max_images = _safe_int(str(payload.get("max_images") or "9"), 9, min_value=1, max_value=20)

    rows = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id).order_by(DatasetItem.row_index.asc()).all()

    job = Job(system="A", job_type="EXPORT_EXCEL", status="running", dataset_id=dataset_id)
    job.started_at = datetime.utcnow()
    job.set_progress({"total": len(rows), "completed": 0, "failed": 0})
    job.set_meta(
        {
            "template_key": ds.template_key,
            "mode": mode,
            "image_columns": image_columns,
            "max_images": max_images,
        }
    )
    db.session.add(job)
    db.session.commit()

    job_id = job.id

    try:
        import openpyxl  # type: ignore

        base_columns = ds.get_columns() or (_taiyang_default_columns() if (ds.template_key or "") == "taiyang" else [])
        if not base_columns:
            base_columns = _taiyang_default_columns()

        extra_columns: List[str] = []
        if mode == "append":
            extra_columns.extend(["新标题", "新图片", "新SKU图片"])
        if image_columns:
            extra_columns.extend([f"image{i}" for i in range(1, max_images + 1)])

        columns = base_columns + [c for c in extra_columns if c not in base_columns]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(columns)

        completed = 0
        failed = 0

        for it in rows:
            attrs = it.get_attributes() or {}
            raw = attrs.get("raw") if isinstance(attrs, dict) else None
            raw_row = raw if isinstance(raw, dict) else {}

            out: Dict[str, Any] = {c: raw_row.get(c) for c in base_columns}

            chosen_title = (it.new_title or "").strip() or (it.title or "").strip()
            chosen_images = it.get_new_images() or it.get_images()

            if mode == "overwrite":
                if "产品名称" in out and chosen_title:
                    out["产品名称"] = chosen_title
                if "产品图片" in out:
                    out["产品图片"] = ",".join(chosen_images or [])
                if "SKU图片" in out:
                    out["SKU图片"] = (chosen_images[0] if chosen_images else None) or (it.variant_image or "") or out.get("SKU图片")
            else:
                new_images = it.get_new_images()
                out["新标题"] = it.new_title or ""
                out["新图片"] = ",".join(new_images or [])
                out["新SKU图片"] = (new_images[0] if new_images else "") or ""

            if image_columns:
                for idx in range(1, max_images + 1):
                    key = f"image{idx}"
                    out[key] = chosen_images[idx - 1] if chosen_images and (idx - 1) < len(chosen_images) else ""

            ws.append([out.get(c) for c in columns])
            completed += 1

        original_name = f"{ds.name}_export.xlsx" if ds.name else f"{ds.id}_export.xlsx"
        safe_name = secure_filename(original_name) or f"{ds.id}_export.xlsx"

        asset = Asset(system="A", kind="excel", name=original_name, storage="local", dataset_id=dataset_id, job_id=job_id)
        asset.set_meta(
            {
                "source": "dataset_export",
                "dataset_id": dataset_id,
                "template_key": ds.template_key,
                "mode": mode,
                "image_columns": image_columns,
                "max_images": max_images,
            }
        )
        db.session.add(asset)
        db.session.flush()

        upload_root = Path(current_app.config["UPLOAD_FOLDER"]).resolve()
        asset_dir = (upload_root / "assets" / asset.id).resolve()
        asset_dir.mkdir(parents=True, exist_ok=True)

        file_path = (asset_dir / safe_name).resolve()
        wb.save(str(file_path))

        asset.file_path = file_path.relative_to(upload_root).as_posix()
        asset.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        try:
            asset.size_bytes = int(file_path.stat().st_size)
        except Exception:
            asset.size_bytes = None

        job = Job.query.get(job_id)
        if job:
            job.status = "succeeded"
            job.completed_at = datetime.utcnow()
            job.set_progress({"total": len(rows), "completed": completed, "failed": failed, "asset_id": asset.id})
            meta = job.get_meta() or {}
            job.set_meta({**meta, "asset_id": asset.id})

        db.session.commit()

        return success_response(
            {
                "job": job.to_dict() if job else {"id": job_id},
                "asset": asset.to_dict(),
                "download_url": f"/api/assets/{asset.id}/download",
            },
            message="exported",
        )

    except Exception as e:
        logger.error("export_dataset_excel failed: %s", e, exc_info=True)
        db.session.rollback()
        job = Job.query.get(job_id)
        if job:
            job.status = "failed"
            job.error_message = str(e)
            job.completed_at = datetime.utcnow()
            job.set_progress({"total": len(rows), "completed": 0, "failed": len(rows)})
            db.session.commit()
        return error_response("SERVER_ERROR", str(e), 500)


@dataset_bp.route("/<dataset_id>/projects/create", methods=["POST"])
def create_projects_from_dataset(dataset_id: str):
    """
    POST /api/datasets/<dataset_id>/projects/create

    v1: 把 DatasetItem 一键生成/绑定为 Project（先跑通 Excel→Project 桥接入口）

    Body (JSON, optional):
      - item_ids?: string[]  # 不传则全量
      - platform_hint?: string  # e.g. shopee/shein/amazon/tiktok/temu
      - download_material?: true|false (default true)  # 尝试把 SKU图片/首图 下载为 Project Material
    """
    dataset_id = (dataset_id or "").strip()
    ds = Dataset.query.get(dataset_id)
    if not ds:
        return error_response("DATASET_NOT_FOUND", "Dataset not found", 404)

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        payload = {}

    platform_hint = (str(payload.get("platform_hint") or "").strip() or None)
    download_material = payload.get("download_material")
    if download_material is None:
        download_material = True
    download_material = bool(download_material)

    item_ids = payload.get("item_ids")
    if isinstance(item_ids, list) and item_ids:
        selected_ids = [str(x).strip() for x in item_ids if str(x).strip()]
        q = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id, DatasetItem.id.in_(selected_ids))
    else:
        q = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id)

    items = q.order_by(DatasetItem.row_index.asc()).all()
    if not items:
        return error_response("DATASET_ITEMS_NOT_FOUND", "No dataset items", 404)

    job = Job(system="A", job_type="DATASET_TO_PROJECT", status="running", dataset_id=dataset_id)
    job.started_at = datetime.utcnow()
    job.set_progress({"total": len(items), "completed": 0, "failed": 0})
    job.set_meta({"platform_hint": platform_hint, "download_material": download_material})
    db.session.add(job)
    db.session.commit()

    job_id = job.id
    file_service = get_file_service(current_app.config["UPLOAD_FOLDER"])

    results: List[Dict[str, Any]] = []
    created = 0
    skipped = 0
    failed = 0

    def _download_image_as_pil(url: str):
        if not url:
            return None
        try:
            import httpx
            from PIL import Image  # type: ignore
            from io import BytesIO

            with httpx.Client(timeout=20.0, follow_redirects=True) as client:
                res = client.get(url, headers={"User-Agent": "Mozilla/5.0"})
                res.raise_for_status()
                content = res.content

            if not content:
                return None
            if len(content) > 25 * 1024 * 1024:
                return None

            img = Image.open(BytesIO(content))
            img.load()
            return img
        except Exception:
            return None

    for it in items:
        it_id = it.id
        try:
            existing_project_id = (it.project_id or "").strip() or None
            if existing_project_id and Project.query.get(existing_project_id):
                results.append({"item_id": it_id, "project_id": existing_project_id, "created": False})
                skipped += 1
                continue

            idea_prompt = _build_project_idea_prompt_from_item(it, platform_hint=platform_hint)
            project = Project(creation_type="idea", project_type="ecom", idea_prompt=idea_prompt, status="DRAFT")
            db.session.add(project)
            db.session.flush()

            it.project_id = project.id

            if download_material:
                images = it.get_images()
                new_images = it.get_new_images()
                src = (new_images[0] if new_images else "") or (it.variant_image or "").strip() or (images[0] if images else "")
                if src:
                    img = _download_image_as_pil(src)
                    if img is not None:
                        rel = file_service.save_material_image(img, project.id)
                        filename = rel.split("/")[-1]
                        url = file_service.get_file_url(project.id, "materials", filename)
                        material = Material(project_id=project.id, filename=filename, relative_path=rel, url=url)
                        db.session.add(material)

            db.session.commit()

            results.append({"item_id": it_id, "project_id": project.id, "created": True})
            created += 1

        except Exception as e:
            logger.warning("create project failed for item %s: %s", it_id, e, exc_info=True)
            db.session.rollback()
            failed += 1
            row = DatasetItem.query.filter(DatasetItem.dataset_id == dataset_id, DatasetItem.id == it_id).first()
            if row:
                errs = row.get_errors()
                msg = f"create_project_failed: {str(e)}"
                if msg not in errs:
                    errs.append(msg)
                    row.set_errors(errs)
                row.status = "failed"
                db.session.commit()
            results.append({"item_id": it_id, "project_id": None, "created": False, "error": str(e)})

    job = Job.query.get(job_id)
    if job:
        job.status = "succeeded" if failed == 0 else "failed"
        job.completed_at = datetime.utcnow()
        job.set_progress({"total": len(items), "completed": created + skipped, "failed": failed})
        meta = job.get_meta() or {}
        job.set_meta({**meta, "created": created, "skipped": skipped, "failed": failed})
        db.session.commit()

    return success_response(
        {
            "job": job.to_dict() if job else {"id": job_id},
            "created": created,
            "skipped": skipped,
            "failed": failed,
            "results": results,
        },
        message="projects_created",
    )
